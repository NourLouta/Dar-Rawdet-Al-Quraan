# -*- coding: utf-8 -*-
"""
طبقة البيانات — القراءة/الكتابة من Google Sheets.

استراتيجية مزدوجة:
  • القراءة: عبر gspread إذا توفّر حساب الخدمة في st.secrets، وإلا عبر رابط
    CSV العام (gviz) — هذا يجعل القراءة تعمل فورًا حتى قبل إعداد الصلاحيات.
  • الكتابة/التحديث: عبر gspread فقط (يتطلب حساب خدمة + مشاركة الملف معه).

التخزين المؤقت قصير (ttl) لتقليل النداءات مع إبقاء البيانات حديثة.
"""
from __future__ import annotations
import logging
import urllib.parse
from functools import lru_cache

import pandas as pd

from . import config
from . import schema

logger = logging.getLogger(__name__)

# ربط المفتاح المنطقي ← (اسم الورقة، مفتاح الترويسة في schema.HEADERS)
WS_MAP = {
    "lookups":     (config.WS_LOOKUPS,         None),
    "teachers":    (config.WS_TEACHERS,        "teachers"),
    "parents":     (config.WS_PARENTS,         "parents"),
    "students":    (config.WS_STUDENTS,        "students"),
    "enrollments": (config.WS_ENROLLMENTS,     "enrollments"),
    "sessions":    (config.WS_SESSIONS,        "sessions"),
    "pfeedback":   (config.WS_PARENT_FEEDBACK, "pfeedback"),
    "tfeedback":   (config.WS_TEACHER_FEEDBACK, None),
}

# عمود المفتاح الأساسي لكل ورقة كيان — تُحذف الصفوف القالبية الفارغة منه عند القراءة
PRIMARY_CODE = {
    "teachers":    schema.Teacher.CODE,
    "parents":     schema.Parent.CODE,
    "students":    schema.Student.CODE,
    "enrollments": schema.Enrollment.CODE,
    "sessions":    schema.Session.CODE,
    "pfeedback":   schema.ParentFeedback.CODE,
}


def _drop_template_rows(key: str, df: pd.DataFrame) -> pd.DataFrame:
    """إزالة الصفوف القالبية الفارغة (بلا كود) التي ينشئها قالب الملف."""
    code_col = PRIMARY_CODE.get(key)
    if code_col and not df.empty and code_col in df.columns:
        mask = df[code_col].astype(str).str.strip().replace({"nan": "", "None": ""}) != ""
        df = df[mask].reset_index(drop=True)
    return df


# ════════════════════════════════════════════════════════════════════════════
# 🔌 الاتصال
# ════════════════════════════════════════════════════════════════════════════
def _secrets():
    try:
        import streamlit as st
        return st.secrets
    except Exception:
        return {}


def get_sheet_id() -> str:
    s = _secrets()
    try:
        return s["sheet"]["id"]
    except Exception:
        return config.SHEET_ID


@lru_cache(maxsize=1)
def get_client():
    """عميل gspread من حساب الخدمة، أو None لو غير متوفّر (وضع القراءة فقط)."""
    s = _secrets()
    try:
        creds_info = dict(s["gcp_service_account"])
    except Exception:
        logger.info("لا يوجد حساب خدمة — وضع القراءة فقط (CSV).")
        return None
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
        return gspread.authorize(creds)
    except Exception as e:
        logger.error(f"فشل تهيئة gspread: {e}")
        return None


def get_script():
    """رابط + رمز Google Apps Script (طريقة الكتابة المجانية بلا حساب خدمة)."""
    s = _secrets()
    try:
        return s["apps_script"]["url"], s["apps_script"].get("token", "")
    except Exception:
        return None, None


def can_write() -> bool:
    """الحفظ متاح عبر: حساب خدمة Google، أو Apps Script، أو ملف xlsx محلي."""
    return write_target() != "none"


def write_target() -> str:
    if get_client() is not None:
        return "google"
    if get_script()[0]:
        return "script"
    if config.LOCAL_XLSX is not None:
        return "local"
    return "none"


@lru_cache(maxsize=1)
def _spreadsheet():
    client = get_client()
    if client is None:
        return None
    try:
        return client.open_by_key(get_sheet_id())
    except Exception as e:
        logger.error(f"تعذّر فتح الملف: {e}")
        return None


# ════════════════════════════════════════════════════════════════════════════
# 📥 القراءة
# ════════════════════════════════════════════════════════════════════════════
def _read_csv(ws_name: str) -> pd.DataFrame:
    """قراءة ورقة عبر رابط gviz العام (قراءة فقط)."""
    sid = get_sheet_id()
    enc = urllib.parse.quote(ws_name)
    url = f"https://docs.google.com/spreadsheets/d/{sid}/gviz/tq?tqx=out:csv&sheet={enc}"
    try:
        df = pd.read_csv(url, encoding="utf-8", dtype=str, keep_default_na=False)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")
        df = df[~(df == "").all(axis=1)]
        return df.reset_index(drop=True)
    except Exception as e:
        logger.warning(f"تعذّر قراءة '{ws_name}' عبر CSV: {e}")
        return pd.DataFrame()


def _read_gspread(ws_name: str) -> pd.DataFrame:
    ss = _spreadsheet()
    if ss is None:
        return _read_csv(ws_name)
    try:
        ws = ss.worksheet(ws_name)
        # numericise_ignore=['all'] يبقي كل القيم نصًّا (يحفظ الأصفار البادئة في الهواتف والأكواد)
        records = ws.get_all_records(numericise_ignore=["all"])
        df = pd.DataFrame(records)
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception as e:
        logger.warning(f"تعذّر قراءة '{ws_name}' عبر gspread: {e} — تجربة CSV.")
        return _read_csv(ws_name)


def _read_local_xlsx(ws_name: str) -> pd.DataFrame:
    """قراءة ورقة من ملف xlsx محلي (وضع التجربة دون اتصال)."""
    path = config.LOCAL_XLSX
    if not path:
        return pd.DataFrame()
    try:
        df = pd.read_excel(path, sheet_name=ws_name, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all").fillna("")
        return df.reset_index(drop=True)
    except Exception as e:
        logger.warning(f"تعذّر قراءة '{ws_name}' من الملف المحلي: {e}")
        return pd.DataFrame()


def read_ws(key: str) -> pd.DataFrame:
    """
    قراءة ورقة واحدة بالمفتاح المنطقي.
    الأولوية: gspread (لو متوفّر) ← ملف xlsx محلي (لو مضبوط) ← رابط CSV العام.
    """
    ws_name = WS_MAP.get(key, (key, None))[0]
    if get_client() is not None:
        df = _read_gspread(ws_name)
    elif config.LOCAL_XLSX is not None:
        df = _read_local_xlsx(ws_name)
    else:
        df = _read_csv(ws_name)
    return _drop_template_rows(key, df)


def load_all(ttl: int = 120) -> dict:
    """قراءة كل الأوراق. يُغلّف بالتخزين المؤقت داخل التطبيق."""
    try:
        import streamlit as st

        @st.cache_data(ttl=ttl, show_spinner=False)
        def _cached():
            return {k: read_ws(k) for k in WS_MAP}
        return _cached()
    except Exception:
        return {k: read_ws(k) for k in WS_MAP}


def clear_cache():
    try:
        import streamlit as st
        st.cache_data.clear()
    except Exception:
        pass
    get_client.cache_clear()
    _spreadsheet.cache_clear()


# ════════════════════════════════════════════════════════════════════════════
# 📋 القوائم المرجعية
# ════════════════════════════════════════════════════════════════════════════
def get_lookups(lookups_df: pd.DataFrame | None = None) -> dict:
    """تحويل ورقة القوائم المرجعية إلى dict: مفتاح → قائمة قيم نظيفة."""
    if lookups_df is None:
        lookups_df = read_ws("lookups")
    out = {}
    if lookups_df is None or lookups_df.empty:
        return out
    for key, col in schema.LOOKUP_COLS.items():
        if col in lookups_df.columns:
            vals = [str(v).strip() for v in lookups_df[col].tolist()
                    if str(v).strip() and str(v).strip().lower() != "nan"]
            out[key] = list(dict.fromkeys(vals))  # إزالة التكرار مع الحفاظ على الترتيب
        else:
            out[key] = []
    return out


# ════════════════════════════════════════════════════════════════════════════
# 🔢 توليد الأكواد
# ════════════════════════════════════════════════════════════════════════════
def next_code(kind: str, existing_df: pd.DataFrame, code_col: str) -> str:
    """توليد الكود التالي (مثل S-00007) بناءً على أكبر رقم موجود."""
    import re
    prefix, width = config.CODE_PREFIX[kind]
    max_n = 0
    if existing_df is not None and not existing_df.empty and code_col in existing_df.columns:
        for v in existing_df[code_col].astype(str):
            v = v.strip()
            if v.startswith(prefix):
                # استخراج الأرقام الأولى بعد البادئة (يتحمّل صيغ العرض "S-7 — اسم")
                m = re.match(r"\d+", schema.normalize_digits(v[len(prefix):]))
                if m:
                    max_n = max(max_n, int(m.group()))
    return f"{prefix}{max_n + 1:0{width}d}"


# ════════════════════════════════════════════════════════════════════════════
# 📤 الكتابة — عبر gspread (Google) أو ملف xlsx محلي (نفس واجهة الاستدعاء)
# ════════════════════════════════════════════════════════════════════════════
class WriteError(RuntimeError):
    pass


# ── gspread ──────────────────────────────────────────────────────────────────
def _require_ws(key: str):
    ss = _spreadsheet()
    if ss is None:
        raise WriteError("الكتابة عبر Google تتطلب حساب خدمة مُعدًّا في الإعدادات.")
    ws_name, hdr_key = WS_MAP[key]
    try:
        return ss.worksheet(ws_name)
    except Exception:
        headers = schema.HEADERS.get(hdr_key, [])
        ws = ss.add_worksheet(title=ws_name, rows=200, cols=max(10, len(headers) + 2))
        if headers:
            ws.append_row(headers, value_input_option="USER_ENTERED")
        return ws


def _header_row(ws) -> list[str]:
    return [str(c).strip() for c in ws.row_values(1)]


# ── ملف محلي (openpyxl) ──────────────────────────────────────────────────────
def _local_ws(wb, key: str):
    """إرجاع ورقة من المصنّف المحلي، وإنشاؤها بالترويسة إن لم تكن موجودة."""
    ws_name, hdr_key = WS_MAP[key]
    if ws_name in wb.sheetnames:
        return wb[ws_name]
    ws = wb.create_sheet(title=ws_name)
    for c, h in enumerate(schema.HEADERS.get(hdr_key, []), start=1):
        ws.cell(row=1, column=c, value=h)
    return ws


def _local_headers(ws, hdr_key) -> list[str]:
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    headers = [h for h in headers if h]
    if not headers:
        headers = schema.HEADERS.get(hdr_key, [])
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c, value=h)
    return headers


# ── Google Apps Script (كتابة مجانية عبر HTTP) ───────────────────────────────
def _script_post(payload: dict) -> dict:
    import json
    import urllib.request
    url, token = get_script()
    if not url:
        raise WriteError("لا يوجد رابط Apps Script في الإعدادات.")
    payload = dict(payload)
    payload["token"] = token or ""
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data,
                                 headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = resp.read().decode("utf-8")
    except Exception as e:
        raise WriteError(f"تعذّر الاتصال بـ Apps Script: {e}")
    try:
        res = json.loads(body)
    except Exception:
        raise WriteError("رد غير متوقع من Apps Script (تحقق من النشر والصلاحيات).")
    if not res.get("ok"):
        raise WriteError(f"Apps Script: {res.get('error')}")
    return res


def _full_headers(key: str, existing: list[str], sample_keys) -> list[str]:
    """دمج الأعمدة الموجودة مع أعمدة المخطط ومفاتيح البيانات (تُضاف الناقصة في النهاية)."""
    out = [h for h in existing if h]
    for h in list(schema.HEADERS.get(WS_MAP[key][1], [])) + list(sample_keys):
        if h and h not in out:
            out.append(h)
    return out


def _local_append(key: str, rows: list[dict]) -> int:
    import openpyxl
    path = config.LOCAL_XLSX
    wb = openpyxl.load_workbook(path)
    ws = _local_ws(wb, key)
    existing = _local_headers(ws, WS_MAP[key][1])
    headers = _full_headers(key, existing, rows[0].keys())
    # كتابة صف الترويسة كاملًا (يضيف الأعمدة الناقصة مثل النمط الأسبوعي)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    for data in rows:
        ws.append([_fmt(data.get(h, "")) for h in headers])
    wb.save(path)
    return len(rows)


def _local_update(key: str, code_col: str, code_val: str, updates: dict) -> bool:
    import openpyxl
    path = config.LOCAL_XLSX
    wb = openpyxl.load_workbook(path)
    ws = _local_ws(wb, key)
    headers = _local_headers(ws, WS_MAP[key][1])
    if code_col not in headers:
        raise WriteError(f"العمود '{code_col}' غير موجود في الورقة.")
    ci = headers.index(code_col)
    found = False
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=ci + 1).value
        if str(cell).strip() == str(code_val).strip():
            for col, val in updates.items():
                if col in headers:
                    ws.cell(row=r, column=headers.index(col) + 1, value=_fmt(val))
            found = True
            break
    if found:
        wb.save(path)
    return found


def _local_delete(key: str, code_col: str, code_val: str) -> bool:
    import openpyxl
    path = config.LOCAL_XLSX
    wb = openpyxl.load_workbook(path)
    ws = _local_ws(wb, key)
    headers = _local_headers(ws, WS_MAP[key][1])
    if code_col not in headers:
        return False
    ci = headers.index(code_col)
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=ci + 1).value).strip() == str(code_val).strip():
            ws.delete_rows(r, 1)
            wb.save(path)
            return True
    return False


# ── واجهة موحّدة ──────────────────────────────────────────────────────────────
def append_row(key: str, data: dict) -> None:
    append_rows(key, [data])


def append_rows(key: str, rows: list[dict]) -> int:
    """إضافة صفوف (Google أو محلي حسب المتاح). يُرجع العدد المضاف."""
    if not rows:
        return 0
    tgt = write_target()
    if tgt == "google":
        ws = _require_ws(key)
        headers = _header_row(ws)
        if not headers:
            headers = schema.HEADERS.get(WS_MAP[key][1], list(rows[0].keys()))
            ws.append_row(headers, value_input_option="USER_ENTERED")
        full = _full_headers(key, headers, rows[0].keys())
        if full != headers:
            # إضافة الأعمدة الناقصة (مثل النمط الأسبوعي) إلى صف الترويسة
            if ws.col_count < len(full):
                ws.add_cols(len(full) - ws.col_count)
            for i, h in enumerate(full):
                if i >= len(headers):
                    ws.update_cell(1, i + 1, h)
            headers = full
        matrix = [[_fmt(r.get(h, "")) for h in headers] for r in rows]
        ws.append_rows(matrix, value_input_option="USER_ENTERED")
        return len(matrix)
    if tgt == "script":
        headers = _full_headers(key, [], rows[0].keys())
        payload_rows = [{h: _fmt(r.get(h, "")) for h in headers} for r in rows]
        res = _script_post({"action": "append", "sheet": WS_MAP[key][0],
                            "headers": headers, "rows": payload_rows})
        return int(res.get("added", len(rows)))
    if tgt == "local":
        return _local_append(key, rows)
    raise WriteError("لا يوجد مصدر كتابة (لا Google ولا Apps Script ولا ملف محلي).")


def update_row_by_code(key: str, code_col: str, code_val: str, updates: dict) -> bool:
    """تحديث صف مطابق للكود (Google أو محلي). يُرجع True عند النجاح."""
    tgt = write_target()
    if tgt == "google":
        ws = _require_ws(key)
        headers = _header_row(ws)
        if code_col not in headers:
            raise WriteError(f"العمود '{code_col}' غير موجود في الورقة.")
        col_idx = headers.index(code_col) + 1
        import re as _re
        cell = ws.find(_re.compile(rf"^{_re.escape(str(code_val))}$"), in_column=col_idx)
        if cell is None:
            return False
        for col, val in updates.items():
            if col in headers:
                ws.update_cell(cell.row, headers.index(col) + 1, _fmt(val))
        return True
    if tgt == "script":
        res = _script_post({"action": "update", "sheet": WS_MAP[key][0],
                            "codeCol": code_col, "codeVal": str(code_val),
                            "updates": {k: _fmt(v) for k, v in updates.items()}})
        return bool(res.get("updated", False))
    if tgt == "local":
        return _local_update(key, code_col, code_val, updates)
    raise WriteError("لا يوجد مصدر كتابة.")


def delete_row_by_code(key: str, code_col: str, code_val: str) -> bool:
    """حذف صف مطابق للكود (Google أو محلي). يُرجع True عند النجاح."""
    tgt = write_target()
    if tgt == "google":
        ws = _require_ws(key)
        headers = _header_row(ws)
        if code_col not in headers:
            return False
        import re as _re
        cell = ws.find(_re.compile(rf"^{_re.escape(str(code_val))}$"),
                       in_column=headers.index(code_col) + 1)
        if cell is None:
            return False
        ws.delete_rows(cell.row)
        return True
    if tgt == "script":
        res = _script_post({"action": "delete", "sheet": WS_MAP[key][0],
                            "codeCol": code_col, "codeVal": str(code_val)})
        return bool(res.get("deleted", False))
    if tgt == "local":
        return _local_delete(key, code_col, code_val)
    raise WriteError("لا يوجد مصدر كتابة.")


def _fmt(v) -> str:
    """تحويل القيمة إلى نص مناسب للكتابة في الخلية."""
    import datetime as _dt
    if v is None:
        return ""
    if isinstance(v, (_dt.date, _dt.datetime)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)
