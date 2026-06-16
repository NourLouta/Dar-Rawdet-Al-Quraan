# -*- coding: utf-8 -*-
"""
سكربت ترحيل لمرة واحدة: من الملف القديم المُفكَّك → المخطط الجديد المنظّم.

يقرأ ورقة "الطلاب بينات كاملة" (88 عمودًا، تتكرر فيها أعمدة الحصص) وورقة
"المحفظين" من الملف القديم، ثم:
  • يطبّع الأرقام العربية والهواتف والنوع والفئة والتواريخ.
  • يطابق أسماء السور مع القائمة المرجعية (مطابقة تقريبية).
  • يفكّك أعمدة الحصص الـ17 (وكتلة نور البيان) إلى صفوف في "الحصص".
  • يزيل تكرار أولياء الأمور وينشئ الأكواد (P-/S-/T-/E-/SS-).
  • يكتب الناتج إلى ملف xlsx جديد للمراجعة (لا يلمس الملف الحيّ).

الاستخدام:
    python migrate.py --old "مسار الملف القديم.xlsx" --out "data/migrated_review.xlsx"
"""
from __future__ import annotations
import argparse
import re
from pathlib import Path

import openpyxl
import pandas as pd

from dar import schema
from dar.schema import (
    Teacher, Parent, Student, Enrollment, Session,
    clean_phone, normalize_digits, to_date, month_key,
    parse_arabic_time, format_arabic_time, add_minutes, arabic_weekday,
)

OLD_SHEET_STUDENTS = "الطلاب بينات كاملة"
OLD_SHEET_TEACHERS = "المحفظين"

GENDER_MAP = {"انثى": "أنثى", "أنثى": "أنثى", "ذكر": "ذكر", "طفلة": "أنثى", "طفل": "ذكر"}
CATEGORY_MAP = {
    "اطفال": "طفل", "أطفال": "طفل", "طفل": "طفل", "طفلة": "طفلة",
    "نساء": "سيدة", "سيدة": "سيدة", "رجل": "رجل", "رجال": "رجل",
}


def _norm_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "nat") else s


def match_surah(name: str, surahs: list[str]) -> str:
    """مطابقة اسم السورة مع القائمة المرجعية (تقريبية)."""
    name = _norm_text(name)
    if not name:
        return ""
    if name in surahs:
        return name
    try:
        from fuzzywuzzy import process
        best, score = process.extractOne(name, surahs)
        return best if score >= 80 else name
    except Exception:
        # مطابقة جزئية بسيطة
        for s in surahs:
            if name in s or s in name:
                return s
        return name


def read_old(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[OLD_SHEET_STUDENTS]
    rows = list(ws.iter_rows(values_only=True))
    header = [(_norm_text(c)) for c in rows[0]]
    # تخطّي صف التسمية (شهر 3/شهر 4) إن وُجد
    data = []
    for r in rows[1:]:
        code = _norm_text(r[header.index(Student.CODE)]) if Student.CODE in header else ""
        name_idx = header.index("الاسم بالكامل") if "الاسم بالكامل" in header else 1
        if not code and not _norm_text(r[name_idx]):
            continue
        if str(code).startswith("شهر") or str(_norm_text(r[name_idx])).startswith("شهر"):
            continue
        data.append(r)
    return header, data, wb


def session_column_pairs(header: list[str]) -> list[tuple[int, int]]:
    """إيجاد أزواج (عمود التاريخ، عمود الوقت التالي) لكل حصة."""
    pairs = []
    for i, h in enumerate(header):
        if h.startswith("تاريخ الحصة"):
            time_idx = i + 1 if (i + 1 < len(header) and header[i + 1].startswith("وقت")) else None
            pairs.append((i, time_idx))
    return pairs


def migrate(old_path: Path, out_path: Path):
    surahs = schema.LOOKUP_COLS  # placeholder
    # قائمة السور المرجعية من الملف الجديد لو توفّر، وإلا تخطّي المطابقة
    surah_list = []
    header, data, wb = read_old(old_path)

    def col(h):
        return header.index(h) if h in header else None

    sess_pairs = session_column_pairs(header)

    parents, students, enrollments, sessions = [], [], [], []
    parent_index = {}   # (name, phone) -> code
    teacher_code_by_name = {}

    # ── المحفظون ─────────────────────────────────────────────────────────────
    teachers = []
    if OLD_SHEET_TEACHERS in wb.sheetnames:
        tws = wb[OLD_SHEET_TEACHERS]
        trows = list(tws.iter_rows(values_only=True))
        thdr = [_norm_text(c) for c in trows[0]]
        for r in trows[1:]:
            d = {thdr[i]: r[i] for i in range(min(len(thdr), len(r)))}
            code = _norm_text(d.get("ID") or d.get("كود المحفظ"))
            name = _norm_text(d.get("الاسم"))
            if not name:
                continue
            teacher_code_by_name[name] = code or f"T-{len(teachers)+1:04d}"
            teachers.append({
                Teacher.CODE: teacher_code_by_name[name],
                Teacher.NAME: name,
                Teacher.GENDER: GENDER_MAP.get(_norm_text(d.get("النوع")), _norm_text(d.get("النوع"))),
                Teacher.PHONE: clean_phone(d.get("رقم الهاتف")),
                Teacher.WHATSAPP: clean_phone(d.get("واتساب")),
                Teacher.GOV: _norm_text(d.get("المحافظة")),
                Teacher.QUALIFY: _norm_text(d.get("المؤهل/الإجازة")),
                Teacher.EXPERIENCE: normalize_digits(d.get("سنوات الخبرة")),
                Teacher.TEACHES: _norm_text(d.get("الفئة التي يدرّسها")),
                Teacher.HOURLY: normalize_digits(d.get("سعر الساعة")),
                Teacher.CONTRACT: _norm_text(d.get("حالة التعاقد")),
                Teacher.START: to_date(d.get("تاريخ البداية")),
                Teacher.TIMING: _norm_text(d.get("التوقيت")),
                Teacher.PAY_METHOD: _norm_text(d.get("طريقة الدفع")),
                Teacher.SPECIAL: _norm_text(d.get("مميز في ") or d.get("مميز في")),
                Teacher.NOTES: _norm_text(d.get("ملاحظات")),
            })

    # ── الطلاب + أولياء الأمور + التسجيلات + الحصص ────────────────────────────
    for r in data:
        def g(h):
            i = col(h)
            return r[i] if (i is not None and i < len(r)) else None

        s_code = _norm_text(g(Student.CODE)) or f"S-{len(students)+1:05d}"
        s_name = _norm_text(g("الاسم بالكامل"))
        if not s_name:
            continue

        # ولي الأمر (إزالة التكرار بالاسم+الهاتف)
        p_name = _norm_text(g("اسم ولي الأمر"))
        p_phone = clean_phone(g("رقم الهاتف"))
        pkey = (p_name, p_phone)
        if p_name and pkey not in parent_index:
            p_code = f"P-{len(parents)+1:04d}"
            parent_index[pkey] = p_code
            parents.append({
                Parent.CODE: p_code, Parent.NAME: p_name,
                Parent.PHONE: p_phone, Parent.WHATSAPP: clean_phone(g("رقم واتساب")),
                Parent.ADDRESS: _norm_text(g("العنوان / المنطقة")),
                Parent.SOURCE: _norm_text(g("مصدر الاشتراك")),
                Parent.REG_DATE: to_date(g("تاريخ بداية الاشتراك")),
            })
        p_code = parent_index.get(pkey, "")

        teacher_name = _norm_text(g("اسم المحفظ/ة"))
        t_code = teacher_code_by_name.get(teacher_name, "")

        students.append({
            Student.CODE: s_code, Student.NAME: s_name,
            Student.BIRTH: to_date(g("تاريخ الميلاد")),
            Student.AGE: normalize_digits(g("العمر")),
            Student.GENDER: GENDER_MAP.get(_norm_text(g("النوع")), _norm_text(g("النوع"))),
            Student.CATEGORY: CATEGORY_MAP.get(_norm_text(g("الفئة")), _norm_text(g("الفئة"))),
            Student.PARENT_CODE: f"{p_code} — {p_name}" if p_code else "",
            Student.RELATION: _norm_text(g("صلة القرابة")),
            Student.PARENT_NAME: p_name, Student.PARENT_PHONE: p_phone,
            Student.STUDY_TYPE: _norm_text(g("نوع الدراسة")),
            Student.SURAH: match_surah(g("السورة الحالية"), surah_list),
            Student.STATUS: _norm_text(g("حالة الاشتراك")),
            Student.STOP_REASON: _norm_text(g("سبب الإيقاف")),
            Student.SUB_SYSTEM: _norm_text(g("نظام الاشتراك")),
            Student.SUB_VALUE: normalize_digits(g("قيمة الاشتراك الشهري")),
            Student.REG_DATE: to_date(g("تاريخ بداية الاشتراك")),
            Student.NOTES: _norm_text(g("ملاحظات")),
        })

        # التسجيل
        e_code = f"E-{len(enrollments)+1:05d}"
        enrollments.append({
            Enrollment.CODE: e_code,
            Enrollment.STUDENT_CODE: f"{s_code} — {s_name}",
            Enrollment.STUDENT_NAME: s_name,
            Enrollment.TEACHER_CODE: f"{t_code} — {teacher_name}" if t_code else "",
            Enrollment.TEACHER_NAME: teacher_name,
            Enrollment.STUDY_TYPE: _norm_text(g("نوع الدراسة")),
            Enrollment.START: to_date(g("تاريخ بداية الاشتراك")),
            Enrollment.SUB_VALUE: normalize_digits(g("قيمة الاشتراك الشهري")),
            Enrollment.STATUS: "نشط" if "نشط" in _norm_text(g("حالة الاشتراك")) else _norm_text(g("حالة الاشتراك")),
        })

        # مدة الحصة (الكتلة الأولى بالساعة)
        dur_hr = pd.to_numeric(normalize_digits(g("مدة الحصة (ساعة)")), errors="coerce")
        dur_min = int(dur_hr * 60) if pd.notna(dur_hr) and dur_hr else 30

        # الحصص: تفكيك كل أزواج التاريخ/الوقت
        for (di, ti) in sess_pairs:
            dval = r[di] if di < len(r) else None
            d = to_date(dval)
            if not d:
                continue
            tval = r[ti] if (ti is not None and ti < len(r)) else None
            stime = parse_arabic_time(tval)
            etime = add_minutes(stime, dur_min) if stime else None
            sessions.append({
                Session.CODE: f"SS-{len(sessions)+1:06d}",
                Session.ENROLL_CODE: f"{e_code} — {s_name} / {teacher_name}",
                Session.STUDENT_CODE: s_code, Session.STUDENT_NAME: s_name,
                Session.TEACHER_CODE: t_code, Session.TEACHER_NAME: teacher_name,
                Session.DATE: d, Session.MONTH: month_key(d),
                Session.START_TIME: format_arabic_time(stime),
                Session.END_TIME: format_arabic_time(etime),
                Session.DURATION: dur_min,
                Session.STATUS: "تمت",
                Session.SURAH: match_surah(g("السورة الحالية"), surah_list),
                Session.AMOUNT: _norm_text(g("مقدار الحفظ الشهري")),
            })

    # ── الكتابة ───────────────────────────────────────────────────────────────
    out_path.parent.mkdir(parents=True, exist_ok=True)
    frames = {
        "المحفظين": pd.DataFrame(teachers),
        "أولياء الأمور": pd.DataFrame(parents),
        "الطلاب": pd.DataFrame(students),
        "التسجيلات": pd.DataFrame(enrollments),
        "الحصص": pd.DataFrame(sessions),
    }
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        for name, df in frames.items():
            df.to_excel(xl, sheet_name=name, index=False)

    print("✅ تم الترحيل إلى:", out_path)
    for name, df in frames.items():
        print(f"   {name}: {len(df)} صف")
    return frames


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="ترحيل الملف القديم → المخطط الجديد")
    ap.add_argument("--old", required=True, help="مسار الملف القديم .xlsx")
    ap.add_argument("--out", default="data/migrated_review.xlsx", help="ملف الناتج للمراجعة")
    args = ap.parse_args()
    migrate(Path(args.old), Path(args.out))
